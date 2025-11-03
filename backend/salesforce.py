import io
import pandas as pd
from fastapi import APIRouter, UploadFile, File, HTTPException
from starlette.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

router = APIRouter()

def extrair_facebook(valor):
    if isinstance(valor, str) and "facebook.com" in valor:
        return valor
    return ""

def extrair_instagram(valor):
    if isinstance(valor, str) and ("instagram.com" in valor or "@") and "facebook.com" not in valor:
        return valor
    return ""

@router.post("/salesforce")
async def converter_planilha_salesforce(
    file: UploadFile = File(...),
):
    filename_lower = file.filename.lower()
    if not filename_lower.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Formato não suportado. Use .xlsx ou .xls.")

    conteudo = await file.read()
    buffer = io.BytesIO(conteudo)
    
    try:
        df = pd.read_excel(buffer, dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler planilha: {str(e)}")

    columns = [
        "Company", "LastName", "MobilePhone", "Email", "Website", "Documento__c",
        "Faturamento_Mensal_N_mero_Exato__c", "Canal_Origem__c", "Segmento__c",
        "Facebook__c", "Instagram__c", "Biblioteca_de_An_ncios__c", "LinkedIn__c",
        "Observa_es_fixadas__c", "Quantidade_de_an_ncios__c", "Email_do_Investidor_que_Indicou__c",
        "Contato_1_Nome__c", "Contato_1_Cargo__c", "Contato_1_Telefone__c",
        "Contato_1_Telefone_2__c", "Contato_1_Telefone_3__c",
        "Contato_2_Nome__c", "Contato_2_Cargo__c", "Contato_2_Telefone__c",
        "Contato_2_Telefone_2__c", "Contato_2_Telefone_3__c",
        "Contato_3_Nome__c", "Contato_3_Cargo__c", "Contato_3_Telefone__c",
        "Contato_3_Telefone_2__c", "Contato_3_Telefone_3__c",
        "Contato_4_Nome__c", "Contato_4_Cargo__c", "Contato_4_Telefone__c",
        "Contato_4_Telefone_2__c", "Contato_4_Telefone_3__c",
    ]

    new_df = pd.DataFrame(columns=columns)

    new_df["Company"] = df.get("Nome do Lead", "")
    new_df["LastName"] = df.get("SOCIO1Nome", "")
    new_df["MobilePhone"] = df.get("SOCIO1Celular1", "")
    new_df["Email"] = df.get("E-mails Válidos de Decisores", "")
    new_df["Website"] = df.get("Site", "")
    new_df["Documento__c"] = df.get("CNPJ", "")
    new_df["Observa_es_fixadas__c"] = df.get("Observação", "")
    new_df["Facebook__c"] = df.get("Rede Social", "").apply(extrair_facebook)
    new_df["Instagram__c"] = df.get("Rede Social", "").apply(extrair_instagram)
    new_df["LinkedIn__c"] = df.get("SOCIO1Linkedin", "")
    new_df["Contato_1_Nome__c"] = df.get("SOCIO1Nome", "")
    new_df["Contato_1_Telefone__c"] = df.get("SOCIO1Celular1", "")
    new_df["Contato_1_Telefone_2__c"] = df.get("SOCIO1Celular2", "")
    new_df["Contato_2_Nome__c"] = df.get("SOCIO2Nome", "")
    new_df["Contato_2_Telefone__c"] = df.get("SOCIO2Celular1", "")
    new_df["Contato_2_Telefone_2__c"] = df.get("SOCIO2Celular2", "")
    new_df["Contato_3_Nome__c"] = df.get("SOCIO3Nome", "")
    new_df["Contato_3_Telefone__c"] = df.get("SOCIO3Celular1", "")
    new_df["Contato_3_Telefone_2__c"] = df.get("SOCIO3Celular2", "")

    excel_buffer = io.BytesIO()
    new_df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)

    wb_original = load_workbook(buffer)
    ws_original = wb_original.active
    wb_novo = load_workbook(excel_buffer)
    ws_novo = wb_novo.active

    mapeamento_cores = {
        "SOCIO1Celular1": "Contato_1_Telefone__c",
        "SOCIO1Celular2": "Contato_1_Telefone_2__c",
        "SOCIO2Celular1": "Contato_2_Telefone__c",
        "SOCIO2Celular2": "Contato_2_Telefone_2__c",
        "SOCIO3Celular1": "Contato_3_Telefone__c",
        "SOCIO3Celular2": "Contato_3_Telefone_2__c",
    }

    colunas_originais = {cell.value: idx+1 for idx, cell in enumerate(ws_original[1])}
    colunas_novas = {cell.value: idx+1 for idx, cell in enumerate(ws_novo[1])}

    for col_orig, col_novo in mapeamento_cores.items():
        if col_orig not in colunas_originais or col_novo not in colunas_novas:
            continue
        idx_orig = colunas_originais[col_orig]
        idx_novo = colunas_novas[col_novo]
        for i in range(2, ws_original.max_row + 1):
            cor = ws_original.cell(row=i, column=idx_orig).fill.start_color.rgb
            if cor and cor != "00000000":
                ws_novo.cell(row=i, column=idx_novo).fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")

    # Salva XLSX final em memória
    final_buffer = io.BytesIO()
    wb_novo.save(final_buffer)
    final_buffer.seek(0)

    return StreamingResponse(
        final_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Salesforce.xlsx"}
    )
